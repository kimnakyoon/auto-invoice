"""주문일이 이틀 지나도록 안 나간 주문에 공급사 사이트로 1:1 문의를 남긴다.

왜 필요한가: 송장조회 결과 엑셀의 '주문일지연' 시트에 '2일 지남' 건이 쌓이면
사람이 그 주문의 공급사 화면을 하나씩 열어 "○○○ 배송 언제 시작하나요?"를
남겨왔다. 그 일을 GUI의 [문의] 버튼 하나로 대신한다.

무엇을 문의하나 - 딱 '2일 지남'으로 기록된 건만이다 (사용자 기준):
  - 3일 이상 지난 건은 하지 않는다. '2일 지남'이던 날에 이미 문의를 남겼을
    테니 또 남기면 같은 주문에 문의가 두 번 쌓인다.
  - '출고예정/도착예정'이 적혀 있어도 2일 지남이면 남긴다 - 예정일이 있다고
    실제로 나가는 건 아니어서다.
  - 기준은 조회를 돌린 날의 결과 엑셀에 적힌 '지난 일수'다. 오늘 다시 계산하지
    않는다 - 엑셀에 '2일'이라고 적힌 그 건이 문의 대상이다. 그 일수는 주말을
    뺀 값이라(order_date.days_since) 금요일 주문은 화요일에 2일이 된다.

같은 주문에 두 번 남기지 않도록 logs/inquiries.json 에 남긴 문의를 적어둔다.
버튼을 하루에 두 번 누르거나, 같은 엑셀로 다시 돌려도 이미 남긴 주문은 건너뛴다.

사이트마다 문의 화면이 달라서 어댑터에 post_inquiry(context, product_url,
recipient_name, headless)가 있는 사이트만 처리하고(지금은 롯데온), 없는
사이트는 '아직 지원 안 함'으로 결과에 남긴다 - 사람이 그 건은 직접 남긴다.
어댑터에 prepare_inquiries(context, product_urls, headless)가 더 있으면 그
사이트의 첫 문의 전에 한 번 불러 배치를 미리 훑게 한다(롯데온은 주문목록
API로 문의 화면 주소를 읽어 주문마다 상세를 여는 일을 던다).
"""

from __future__ import annotations

import contextlib
import json
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from . import browser as browser_mod
from . import rate_limit
from .config import load_settings
from .report import LOG_DIR
from .result_excel import DEFAULT_DIR as RESULT_DIR
from .result_excel import STALE_HEADERS, STALE_SHEET_NAME
from .result_excel import (
    _CELL_BORDER,
    _GROUP_LINE,
    _HEADER_FILL,
    _TEXT_FORMAT,
    _THIN,
    _TITLE_FILL,
    _write_summary_band,
)
from .suppliers import common
from .suppliers.base import AdapterError, BlockedError
from .suppliers.registry import get_adapter

# 결과 엑셀 '주문일지연' 시트의 '지난 일수' 칸이 이 값인 건만 문의한다.
TARGET_DAYS_TEXT = "2일"

# 남긴 문의 장부. 같은 마켓 주문번호에 두 번 남기지 않는 근거다.
LEDGER_PATH = LOG_DIR / "inquiries.json"

# 결과 엑셀 파일 이름 (result_excel.save_run_result와 같은 규칙).
RESULT_GLOB = "송장조회결과_*.xlsx"

LogFn = Callable[[str], None]


@dataclass
class InquiryTarget:
    """엑셀 '주문일지연' 시트의 '2일 지남' 한 줄."""

    order_id: str          # 마켓 주문번호 (샵마인 기준)
    recipient_name: str
    product_url: str
    order_date: str = ""
    delivery_note: str = ""

    @property
    def site_key(self) -> str | None:
        adapter = get_adapter(self.product_url) if self.product_url else None
        return getattr(adapter, "SITE_KEY", None) if adapter else None


@dataclass
class InquiryResult:
    order_id: str
    recipient_name: str
    site_key: str
    status: str            # "success" | "fail" | "skip"
    reason: str = ""
    product_url: str = ""
    order_date: str = ""
    delivery_note: str = ""
    message: str = ""      # 실제로 남긴(또는 남기려던) 문의 문구

    @classmethod
    def of(cls, target: "InquiryTarget", site: str | None, status: str, reason: str,
           message: str = "") -> "InquiryResult":
        return cls(target.order_id, target.recipient_name, site or "", status, reason,
                   target.product_url, target.order_date, target.delivery_note, message)


@dataclass
class InquiryRun:
    excel_path: Path | None = None          # 읽은 송장조회 결과 엑셀
    result_excel_path: Path | None = None   # 이번 문의 결과를 적은 엑셀 (바탕화면)
    targets: list[InquiryTarget] = field(default_factory=list)
    results: list[InquiryResult] = field(default_factory=list)
    stopped_reason: str | None = None

    def counts(self) -> dict[str, int]:
        out = {"success": 0, "fail": 0, "skip": 0}
        for r in self.results:
            out[r.status] = out.get(r.status, 0) + 1
        return out


# --------------------------------------------------------------------------
# 결과 엑셀에서 문의할 주문 고르기
# --------------------------------------------------------------------------

def find_latest_result_excel(directory: Path | None = None) -> Path | None:
    """바탕화면에서 가장 최근 '송장조회결과_*.xlsx'. 엑셀이 열어둔 잠금 파일(~$)은 뺀다."""
    directory = directory or RESULT_DIR
    candidates = [p for p in directory.glob(RESULT_GLOB) if not p.name.startswith("~$")]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_targets(excel_path: str | Path) -> list[InquiryTarget]:
    """'주문일지연' 시트에서 '지난 일수'가 2일인 줄만 읽는다.

    칸은 헤더 이름으로 찾는다(result_excel.STALE_HEADERS) - 시트에 칸을 넣고
    빼도 안 깨지게. 시트가 없으면(그날 지연 건이 없었다) 빈 목록이다.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if STALE_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[STALE_SHEET_NAME]
        rows = ws.iter_rows(values_only=True)
        header: list[str] | None = None
        for row in rows:
            if row and "마켓 주문번호" in [str(c).strip() for c in row if c is not None]:
                header = [str(c).strip() if c is not None else "" for c in row]
                break
        if header is None:
            raise ValueError(f"'{STALE_SHEET_NAME}' 시트에서 머리글 줄을 찾지 못했습니다: {excel_path}")
        missing = [h for h in STALE_HEADERS if h in ("마켓 주문번호", "수령인", "지난 일수", "상품URL")
                   and h not in header]
        if missing:
            raise ValueError(f"'{STALE_SHEET_NAME}' 시트에 {missing} 칸이 없습니다: {excel_path}")

        def cell(row, name: str) -> str:
            i = header.index(name) if name in header else -1
            if i < 0 or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        targets: list[InquiryTarget] = []
        seen: set[str] = set()
        for row in rows:
            if not row or cell(row, "마켓 주문번호") == "":
                continue
            if cell(row, "지난 일수") != TARGET_DAYS_TEXT:
                continue
            order_id = cell(row, "마켓 주문번호")
            if order_id in seen:
                continue   # 한 주문이 상품별로 여러 줄이어도 문의는 한 번
            seen.add(order_id)
            targets.append(InquiryTarget(
                order_id=order_id,
                recipient_name=cell(row, "수령인"),
                product_url=cell(row, "상품URL"),
                order_date=cell(row, "주문일"),
                delivery_note=cell(row, "출고/도착예정"),
            ))
        return targets
    finally:
        wb.close()


# --------------------------------------------------------------------------
# 남긴 문의 장부
# --------------------------------------------------------------------------

def load_ledger(path: Path = LEDGER_PATH) -> list[dict]:
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return []
    return data if isinstance(data, list) else []


def posted_order_ids(ledger: list[dict]) -> set[str]:
    return {str(e.get("order_id", "")) for e in ledger if e.get("order_id")}


def _append_ledger(entry: dict, path: Path = LEDGER_PATH) -> None:
    """한 건 남길 때마다 바로 적는다 - 도중에 멈춰도 남긴 건은 장부에 있어야 한다."""
    ledger = load_ledger(path)
    ledger.append(entry)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2), encoding="utf-8")


# --------------------------------------------------------------------------
# 실행
# --------------------------------------------------------------------------

def plan(excel_path: str | Path | None = None
         ) -> tuple[Path | None, list[InquiryTarget], dict[str, list[InquiryTarget]], list[InquiryResult]]:
    """무엇을 문의할지 정한다 (실제로 남기기 전에 GUI가 확인창에 보여주는 용도).

    돌려주는 것: (엑셀 경로, 2일 지남 전체, 사이트별 문의할 건, 문의하지 않고
    넘기는 건의 결과) - 넘기는 건은 이미 남긴 주문, 지원하지 않는 사이트,
    수령인/URL이 빈 줄이다.
    """
    path = Path(excel_path) if excel_path else find_latest_result_excel()
    if path is None or not path.exists():
        return None, [], {}, []
    targets = load_targets(path)
    already = posted_order_ids(load_ledger())
    by_site: dict[str, list[InquiryTarget]] = {}
    skipped: list[InquiryResult] = []
    for t in targets:
        adapter = get_adapter(t.product_url) if t.product_url else None
        site = getattr(adapter, "SITE_KEY", None)
        if t.order_id in already:
            why = "이미 문의를 남긴 주문"
        elif adapter is None or site is None:
            why = "상품URL이 없거나 모르는 사이트"
        elif getattr(adapter, "post_inquiry", None) is None:
            why = f"{site}는 아직 문의 자동화를 지원하지 않음 - 직접 남겨주세요"
        elif not t.recipient_name:
            why = "수령인 이름이 비어 있어 문의 문구를 만들 수 없음"
        else:
            by_site.setdefault(site, []).append(t)
            continue
        skipped.append(InquiryResult.of(t, site, "skip", why))
    return path, targets, by_site, skipped


def run(excel_path: str | Path | None = None, *, limit: int | None = None,
        headless: bool = False, dry_run: bool = False, log: LogFn = print) -> InquiryRun:
    """결과 엑셀의 '2일 지남' 건에 문의를 남긴다.

    사이트별로 브라우저 하나를 열어 한 건씩 순서대로 남긴다(같은 사이트에
    몰아치지 않도록 송장조회와 같은 요청 간격을 지킨다). 한 사이트가 로그인
    등으로 막히면 그 사이트의 나머지는 바로 넘기고 다른 사이트는 계속한다.
    dry_run이면 무엇을 남길지만 보여주고 아무것도 남기지 않는다.
    """
    result = InquiryRun()
    path, targets, by_site, skipped = plan(excel_path)
    result.excel_path = path
    result.targets = targets
    result.results.extend(skipped)
    if path is None:
        result.stopped_reason = f"바탕화면에 '{RESULT_GLOB}' 파일이 없습니다. 먼저 송장 조회를 돌려주세요."
        return result

    log(f"결과 엑셀: {path.name}")
    log(f"'{STALE_SHEET_NAME}' 시트의 {TARGET_DAYS_TEXT} 지남 {len(targets)}건 중 "
        f"문의할 건 {sum(len(v) for v in by_site.values())}건, 넘기는 건 {len(skipped)}건")
    for r in skipped:
        log(f"  - 넘김 {r.order_id} ({r.recipient_name or '이름 없음'}): {r.reason}")

    if limit is not None:
        remaining = limit
        for site in list(by_site):
            by_site[site] = by_site[site][:max(0, remaining)]
            remaining -= len(by_site[site])
            if not by_site[site]:
                del by_site[site]

    if dry_run:
        for site, items in by_site.items():
            for t in items:
                message = _message_for(t)
                log(f"  [{site}] {t.order_id} {t.recipient_name}: '{message}' (미리보기 - 남기지 않음)")
                result.results.append(InquiryResult.of(t, site, "skip", "미리보기 (dry-run)", message))
        return result

    settings = load_settings()
    for site, items in by_site.items():
        _post_site(site, items, settings=settings, headless=headless, result=result, log=log)
    counts = result.counts()
    log(f"문의 남기기 끝: 성공 {counts['success']} / 실패 {counts['fail']} / 넘김 {counts['skip']}")
    return result


def _message_for(target: InquiryTarget) -> str:
    """어댑터가 문구를 정하면 그것을, 없으면 기본 문구를."""
    make = getattr(get_adapter(target.product_url), "inquiry_message", None)
    return make(target.recipient_name) if make else f"{target.recipient_name} 배송 언제 시작하나요?"


def _post_site(site: str, items: list[InquiryTarget], *, settings, headless: bool,
               result: InquiryRun, log: LogFn) -> None:
    """한 사이트의 주문들을 브라우저 하나로 한 건씩 남긴다.

    요청 간격은 송장조회와 같은 방식이다 - '앞 건을 시작한 시각 + 간격' 전에는
    다음 건을 시작하지 않는다(간격에 조회 시간이 포함된다). 로그인이 막히면
    (BlockedError) 남은 건은 시도하지 않고 바로 넘긴다 - 주문마다 몇 분씩
    로그인 대기를 반복하지 않도록.
    """
    adapter = get_adapter(items[0].product_url)
    started = time.monotonic()
    total = len(items)

    def record(i: int, t: InquiryTarget, status: str, reason: str, message: str) -> None:
        result.results.append(InquiryResult.of(t, site, status, reason, message))
        verb = {"success": f"남김 - '{message}'", "fail": f"실패 - {reason}"}.get(status, reason)
        log(f"[{site}] {i}/{total} {t.order_id} {t.recipient_name}: {verb}")

    with sync_playwright() as p, contextlib.ExitStack() as stack:
        browser, context = browser_mod.get_context(
            p, site, headless=headless,
            context_kwargs=getattr(adapter, "CONTEXT_KWARGS", None))
        stack.callback(browser.close)
        # 세션 저장은 브라우저를 닫기 전에 - ExitStack은 나중에 넣은 것을 먼저 푼다.

        def _save_state() -> None:
            with contextlib.suppress(Exception):
                browser_mod.save_state(context, site)

        stack.callback(_save_state)
        # 어댑터가 배치를 미리 훑을 수 있으면 첫 문의 전에 한 번 (롯데온은 주문목록
        # API로 문의 화면 주소를 읽어 주문마다 상세를 여는 일을 던다).
        prepare = getattr(adapter, "prepare_inquiries", None)
        if prepare is not None:
            prepare(context, [t.product_url for t in items], headless=headless)
        blocked_reason: str | None = None
        next_allowed = 0.0
        for i, t in enumerate(items, start=1):
            message = _message_for(t)
            if blocked_reason is not None:
                record(i, t, "skip", f"앞 주문에서 막혀 건너뜀: {blocked_reason}", message)
                continue
            common.sleep(next_allowed - time.monotonic())
            next_allowed = time.monotonic() + rate_limit.request_gap(
                settings.delay_min, settings.delay_max)
            try:
                done = adapter.post_inquiry(context, t.product_url, t.recipient_name,
                                            headless=headless)
            except Exception as e:  # noqa: BLE001 - 한 건의 오류가 나머지를 막으면 안 된다
                reason = str(e) if isinstance(e, AdapterError) else f"{type(e).__name__}: {e}"
                if isinstance(e, BlockedError):
                    blocked_reason = reason
                record(i, t, "fail", reason, message)
                continue
            _append_ledger({
                "order_id": t.order_id,
                "site": site,
                "recipient_name": t.recipient_name,
                "product_url": t.product_url,
                "order_date": t.order_date,
                "message": message,
                "posted_at": datetime.now().isoformat(timespec="seconds"),
            })
            record(i, t, "success", done, message)
    log(f"[{site}] {total}건에 {time.monotonic() - started:.1f}초 걸렸습니다.")


def summarize(run_result: InquiryRun) -> str:
    counts = run_result.counts()
    lines = [f"문의 남기기 결과: 남김 {counts['success']}건 / 실패 {counts['fail']}건 / 넘김 {counts['skip']}건"]
    fails = [r for r in run_result.results if r.status == "fail"]
    if fails:
        lines.append("")
        lines.append("남기지 못한 주문 (직접 남겨주세요):")
        lines.extend(f"  {r.order_id} ({r.recipient_name}) - {r.reason}" for r in fails)
    unsupported = [r for r in run_result.results
                   if r.status == "skip" and "지원하지 않음" in r.reason]
    if unsupported:
        lines.append("")
        lines.append("아직 자동화하지 않은 사이트 (직접 남겨주세요):")
        lines.extend(f"  {r.order_id} ({r.recipient_name}) - {r.site_key}: {r.product_url}"
                     for r in unsupported)
    if run_result.stopped_reason:
        lines.append("")
        lines.append(run_result.stopped_reason)
    return "\n".join(lines)


def save_run_log(run_result: InquiryRun) -> Path | None:
    """이번 실행의 건별 결과를 logs/inquiry_*.json 으로 남긴다 (장부와는 별개)."""
    if not run_result.results:
        return None
    LOG_DIR.mkdir(exist_ok=True)
    path = LOG_DIR / f"inquiry_{datetime.now():%Y%m%d_%H%M%S}.json"
    payload = {
        "excel": str(run_result.excel_path) if run_result.excel_path else None,
        "results": [asdict(r) for r in run_result.results],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


# --------------------------------------------------------------------------
# 문의 결과 엑셀 (바탕화면)
# --------------------------------------------------------------------------
# 송장조회 결과 엑셀과 같은 생김새로 둔다 - 맨 위 제목 띠와 건수 한 줄, '결과'
# 칸은 색 배지, 줄은 같은 계열 옅은 색. 사람 손이 먼저 가야 하는 순서로
# 정렬한다: 실패(직접 남겨야 한다) -> 넘김 중 미지원 사이트(역시 직접) ->
# 남김 -> 넘김 중 이미 남긴 주문(할 게 없다).
INQUIRY_SHEET_NAME = "문의결과"
INQUIRY_SUMMARY_SHEET_NAME = "요약"
INQUIRY_HEADERS = ["결과", "마켓 주문번호", "수령인", "사이트", "주문일", "출고/도착예정",
                   "문의 내용", "사유 / 완료 문구", "상품URL"]
INQUIRY_WIDTHS = [10, 22, 14, 12, 14, 24, 34, 60, 60]
_INQ_COL_RESULT = INQUIRY_HEADERS.index("결과")
_INQ_COL_ORDER_ID = INQUIRY_HEADERS.index("마켓 주문번호")
_INQ_COL_REASON = INQUIRY_HEADERS.index("사유 / 완료 문구")

STATUS_LABELS = {"success": "남김", "fail": "실패", "skip": "넘김"}
_INQ_COLORS = {
    "실패": ("D93025", "FFFFFF", "FCE8E6"),
    "넘김": ("5F6368", "FFFFFF", "F1F3F4"),
    "남김": ("188038", "FFFFFF", "E6F4EA"),
}
_INQ_ACTIONS = {
    "실패": "문의를 남기지 못함 - 상품URL을 열어 직접 남겨주세요 (문의 내용 칸의 문구 그대로)",
    "넘김": "이미 남긴 주문이거나 아직 자동화하지 않은 사이트 - 사유 칸 참고 (미지원 사이트는 직접 남겨주세요)",
    "남김": "공급사 사이트에 문의가 올라감 - 답변은 문자/알림톡으로 옴",
}
_INQ_HEADER_ROW = 3
_UNSUPPORTED_MARK = "지원하지 않음"


def result_label(r: InquiryResult) -> str:
    return STATUS_LABELS.get(r.status, r.status)


def _inq_sort_key(r: InquiryResult) -> tuple[int, str]:
    if r.status == "fail":
        rank = 0
    elif r.status == "skip" and _UNSUPPORTED_MARK in r.reason:
        rank = 1
    elif r.status == "success":
        rank = 2
    else:
        rank = 3
    return rank, r.site_key


def write_inquiry_excel(run_result: InquiryRun, path: str | Path) -> Path:
    path = Path(path)
    rows = sorted(run_result.results, key=_inq_sort_key)
    wb = Workbook()
    ws = wb.active
    ws.title = INQUIRY_SHEET_NAME
    last_col = len(INQUIRY_HEADERS)
    now = datetime.now()

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    head = ws.cell(row=1, column=1, value=f"문의 결과  {now:%Y-%m-%d %H:%M}   전체 {len(rows)}건")
    head.font = Font(bold=True, size=14, color="FFFFFF")
    head.fill = _TITLE_FILL
    head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    counts = run_result.counts()
    unsupported = sum(1 for r in rows if r.status == "skip" and _UNSUPPORTED_MARK in r.reason)
    line = (f"남김 {counts['success']}건     실패 {counts['fail']}건     넘김 {counts['skip']}건"
            + (f" (그중 아직 지원하지 않는 사이트 {unsupported}건 - 직접 남겨주세요)" if unsupported else ""))
    cell = ws.cell(row=2, column=1, value=line)
    cell.font = Font(bold=True, size=11, color="3C4043")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    ws.append(INQUIRY_HEADERS)
    for c in ws[_INQ_HEADER_ROW]:
        c.font = Font(bold=True, color="202124")
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _CELL_BORDER
    ws.row_dimensions[_INQ_HEADER_ROW].height = 22

    for i, r in enumerate(rows):
        label = result_label(r)
        ws.append([label, r.order_id, r.recipient_name, r.site_key, r.order_date,
                   r.delivery_note, r.message, r.reason, r.product_url])
        row = ws[ws.max_row]
        badge_fill, badge_font, row_fill = _INQ_COLORS.get(label, ("5F6368", "FFFFFF", "FFFFFF"))
        group_end = i + 1 == len(rows) or _inq_sort_key(rows[i + 1])[0] != _inq_sort_key(r)[0]
        border = Border(left=_THIN, right=_THIN, top=_THIN,
                        bottom=_GROUP_LINE if group_end else _THIN)
        for c in row:
            c.fill = PatternFill("solid", fgColor=row_fill)
            c.border = border
            c.alignment = Alignment(vertical="center")
        badge = row[_INQ_COL_RESULT]
        badge.fill = PatternFill("solid", fgColor=badge_fill)
        badge.font = Font(bold=True, color=badge_font)
        badge.alignment = Alignment(horizontal="center", vertical="center")
        row[_INQ_COL_ORDER_ID].number_format = _TEXT_FORMAT
        row[_INQ_COL_REASON].alignment = Alignment(wrap_text=True, vertical="center")

    for i, width in enumerate(INQUIRY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{_INQ_HEADER_ROW + 1}"
    ws.auto_filter.ref = (f"A{_INQ_HEADER_ROW}:{get_column_letter(last_col)}"
                          f"{max(ws.max_row, _INQ_HEADER_ROW)}")

    _write_inquiry_summary(wb.create_sheet(INQUIRY_SUMMARY_SHEET_NAME), run_result, now)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _write_inquiry_summary(ws, run_result: InquiryRun, now: datetime) -> None:
    _write_summary_band(ws, 1, "결과별 건수")
    for col, name in enumerate(("결과", "건수", "해야 할 일"), start=1):
        c = ws.cell(row=2, column=col, value=name)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _CELL_BORDER
    counts = run_result.counts()
    for status in ("fail", "skip", "success"):
        label = STATUS_LABELS[status]
        ws.append([label, counts.get(status, 0), _INQ_ACTIONS[label]])
        row = ws[ws.max_row]
        badge_fill, badge_font, row_fill = _INQ_COLORS[label]
        for c in row[:3]:
            c.fill = PatternFill("solid", fgColor=row_fill)
            c.border = _CELL_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        row[0].fill = PatternFill("solid", fgColor=badge_fill)
        row[0].font = Font(bold=True, color=badge_font)
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].font = Font(bold=True)
        row[1].alignment = Alignment(horizontal="center", vertical="center")

    _write_summary_band(ws, ws.max_row + 2, "실행 정보")
    info = [("실행 시각", f"{now:%Y-%m-%d %H:%M:%S}"),
            ("문의 기준", f"송장조회 결과 엑셀 '{STALE_SHEET_NAME}' 시트의 지난 일수 {TARGET_DAYS_TEXT} 건"),
            ("읽은 결과 엑셀", str(run_result.excel_path or "")),
            ("문의 장부", str(LEDGER_PATH))]
    if run_result.stopped_reason:
        info.append(("멈춘 이유", run_result.stopped_reason))
    for name, value in info:
        ws.append([name, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=3)
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 78


def save_result_excel(run_result: InquiryRun, *, out_dir: Path | None = None,
                      log: LogFn = print) -> Path | None:
    """이번 문의 결과를 바탕화면 '문의결과_시각.xlsx'로 남긴다.

    결과가 한 건도 없으면(엑셀이 없어 시작도 못 함) 만들지 않는다. 저장에
    실패해도 문의 결과 자체를 덮으면 안 되므로 경고 한 줄만 남기고 None.
    """
    if not run_result.results:
        return None
    out_dir = out_dir or RESULT_DIR
    path = out_dir / f"문의결과_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    try:
        saved = write_inquiry_excel(run_result, path)
    except Exception as e:  # noqa: BLE001 - 엑셀 저장 실패가 실행 결과를 덮으면 안 된다
        log(f"경고: 문의 결과 엑셀을 저장하지 못했습니다 - {e}")
        return None
    run_result.result_excel_path = saved
    log(f"문의 결과 엑셀: {saved}")
    return saved
