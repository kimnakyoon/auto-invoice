"""송장조회 결과를 사람이 바로 열어볼 엑셀로 남긴다 (바탕화면).

왜 필요한가: 6~8단계(샵마인 반영)에서 멈춰도 5단계 송장조회는 이미 끝나 있는
경우가 대부분이다. 그때 조회 결과가 logs/run_*.json 에만 남으면 사람이 JSON을
열어봐야 하고, 멈춘 뒤 어느 주문을 직접 처리해야 하는지 한눈에 안 들어온다.
그래서 같은 내용을 바탕화면 엑셀로도 남긴다 - 반영까지 성공한 실행이든,
중간에 멈춘 실행이든 조회를 한 번이라도 했으면 항상 만든다.

정렬은 파일 순서가 아니라 '사람 손이 먼저 가야 하는 순서'로 한다
(실패 -> 취소/품절 -> 반영오류 -> 주문일지연 -> 미지원 사이트 -> 스킵 -> 성공).
성공
건은 이미 업로드용 CSV로 처리되므로 맨 뒤여도 되고, 확인이 필요한 건이 위로
올라온다.

보기 편하라고 넣은 것들: 맨 위에 실행 시각과 결과별 건수 한 줄, '결과' 칸은
색깔 배지, 줄 전체는 같은 계열 옅은 색, 결과가 바뀌는 자리에는 굵은 가로선.
파일을 열자마자 '확인해야 할 게 몇 건인지'가 먼저 보이는 게 목적이다.

아직 송장이 안 나와 넘긴 건 중 주문일이 이틀 이상 지난 건
(report.is_stale_entry)은 '주문일' 칸을 빨갛게 표시하고, 그런 건만 모은
시트를 따로 하나 더 만든다 - 아직 안 나갔는데 주문한 지 며칠 됐다는
뜻이라 목록으로 한 번에 봐야 해서다.

주문일이 오늘과 이틀 이상 벌어진 건에는 '상품URL'을 같이 적는다. 그 건들은
사람이 공급사 화면을 직접 열어 확인하게 되는데, 주문번호만으로 그 화면을 다시
찾아가는 게 번거로워서다. 나머지 건은 빈칸으로 둔다 - 모든 줄에 URL이 있으면
정작 봐야 할 줄이 묻힌다. '출고/도착예정'(eta.py)은 읽힌 건이면 어디든 적는다.
둘은 같이 보는 값이라(예정일을 보고 화면을 열어 확인) 나란히 붙여 둔다.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import order_date as order_date_mod
from .models import ReportEntry
from .report import is_stale_entry, result_label, stale_entries

# 사람이 바로 열어볼 수 있게 바탕화면에 둔다 (내보내기 엑셀/업로드 CSV와 같은 곳).
DEFAULT_DIR = Path.home() / "Desktop"

SHEET_NAME = "송장조회결과"
SUMMARY_SHEET_NAME = "요약"
# 주문일이 오래된 건만 모아 보여주는 시트 (해당 건이 있을 때만 만든다).
STALE_SHEET_NAME = "주문일지연"

HEADERS = ["결과", "마켓 주문번호", "수령인", "주문일", "출고/도착예정", "상품URL",
           "택배사", "송장번호", "샵마인 반영", "사유"]
COLUMN_WIDTHS = [16, 22, 10, 22, 24, 46, 12, 22, 16, 62]

# 칸 위치를 숫자로 세지 않고 헤더 이름으로 잡는다 (컬럼을 넣고 빼도 안 깨진다).
COL_RESULT = HEADERS.index("결과")
COL_ORDER_ID = HEADERS.index("마켓 주문번호")
COL_ORDER_DATE = HEADERS.index("주문일")
COL_ETA = HEADERS.index("출고/도착예정")
COL_TRACKING = HEADERS.index("송장번호")
COL_APPLIED = HEADERS.index("샵마인 반영")
COL_REASON = HEADERS.index("사유")
COL_URL = HEADERS.index("상품URL")

STALE_HEADERS = ["마켓 주문번호", "수령인", "주문일", "지난 일수", "출고/도착예정",
                 "상품URL", "조회 결과", "사유"]
STALE_WIDTHS = [22, 10, 14, 10, 24, 46, 16, 62]
STALE_COL_DAYS = STALE_HEADERS.index("지난 일수")
STALE_COL_RESULT = STALE_HEADERS.index("조회 결과")
STALE_COL_REASON = STALE_HEADERS.index("사유")

# 위에 올라올수록 사람이 먼저 봐야 하는 결과. '주문일지연'은 결과 이름이
# 아니라 주문일로 판정하는 값이라(성공/스킵 건에도 걸린다) 세는 데는 쓰이지
# 않고, 정렬할 때 그 건들을 끌어올릴 자리로만 쓴다.
_STALE_SORT_LABEL = STALE_SHEET_NAME
_SORT_ORDER = ["실패", "취소/품절", "반영오류", _STALE_SORT_LABEL, "미지원 사이트",
               "스킵", "성공"]

# 결과별 색: (배지 바탕, 배지 글씨, 줄 바탕). 배지는 진하게, 줄은 옅게 해서
# 스크롤할 때 색 덩어리만 보고도 어디까지가 같은 결과인지 알 수 있게 한다.
_COLORS = {
    "실패": ("D93025", "FFFFFF", "FCE8E6"),
    "미지원 사이트": ("E8710A", "FFFFFF", "FEF0E0"),
    "취소/품절": ("9334E6", "FFFFFF", "F3E8FD"),
    "반영오류": ("B31412", "FFFFFF", "FAD2CF"),
    "스킵": ("5F6368", "FFFFFF", "F1F3F4"),
    "성공": ("188038", "FFFFFF", "E6F4EA"),
}
_DEFAULT_COLOR = ("5F6368", "FFFFFF", "FFFFFF")

# 주문일이 오래된 건을 눈에 띄게 하는 색 (결과 색과 겹치지 않게 글씨만 빨갛게).
_STALE_FONT = Font(bold=True, color="D93025")

# 결과별로 사람이 뭘 해야 하는지 한 마디 - 요약 시트에서 쓴다.
_ACTIONS = {
    "실패": "직접 확인 필요",
    "미지원 사이트": "공급사 사이트에서 직접 조회 (어댑터 없음)",
    "취소/품절": "취소/품절 주문인지 확인",
    "반영오류": "샵마인 [송장번호수정] 결과 창에 오류로 뜬 주문 - 샵마인에서 직접 확인",
    "스킵": "그냥 두면 되는 건 (송장 미발급 등)",
    "성공": "송장 확보 - 업로드용 CSV로 반영",
}

_TITLE_FILL = PatternFill("solid", fgColor="202124")
_HEADER_FILL = PatternFill("solid", fgColor="E8EAED")
_THIN = Side(style="thin", color="D0D3D7")
_GROUP_LINE = Side(style="medium", color="80868B")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 주문번호/송장번호는 자릿수가 길어 엑셀이 지수표기(2.21E+13)로 바꿔버린다.
_TEXT_FORMAT = "@"

# 제목 2줄(제목/건수) 다음이 헤더, 그 다음부터 데이터.
_HEADER_ROW = 3


def _sort_key(entry: ReportEntry) -> tuple[int, int]:
    """실패 -> 취소/품절 -> 반영오류 -> 주문일지연 -> 미지원 사이트 -> 스킵 -> 성공.

    주문일이 오래된 건은 결과가 무엇이든 주문일지연 자리까지 끌어올린다. 이미
    그보다 위에 있는 결과(실패/취소·품절)는 제자리를 지킨다 - 더 급한 쪽이
    위여야 하니까. 같은 자리 안에서는 결과끼리 다시 모이도록 두 번째 값으로
    결과 순서를 쓴다(예: 주문일지연 칸의 스킵 건과 성공 건이 섞이지 않게).
    """
    label = result_label(entry)
    rank = _SORT_ORDER.index(label) if label in _SORT_ORDER else len(_SORT_ORDER)
    if is_stale_entry(entry):
        return min(rank, _SORT_ORDER.index(_STALE_SORT_LABEL)), rank
    return rank, rank


def label_counts(entries: list[ReportEntry]) -> list[tuple[str, int]]:
    """결과별 건수를 '사람 손이 먼저 가는 순서'로 돌려준다 (0건은 빼고)."""
    counts: dict[str, int] = {}
    for entry in entries:
        label = result_label(entry)
        counts[label] = counts.get(label, 0) + 1
    ordered = [(label, counts.pop(label)) for label in _SORT_ORDER if label in counts]
    ordered.extend(sorted(counts.items()))
    return ordered


def write_result_excel(
    entries: list[ReportEntry],
    path: str | Path,
    *,
    applied_label: str = "",
    summary: list[tuple[str, str]] | None = None,
    title: str = "송장조회 결과",
    apply_errors: list[str] | None = None,
) -> Path:
    """조회 결과 목록을 엑셀로 저장하고 저장한 경로를 돌려준다.

    applied_label: 성공 건의 '샵마인 반영' 칸에 넣을 문구. 반영까지 갔는지는
    주문 단위로 알 수 없고(샵마인 결과 창은 건수만 준다) 실행 단위로만 알 수
    있어서, 실행 결과를 그대로 성공 건에 붙인다.
    summary: 요약 시트의 '실행 정보'에 넣을 (항목, 값) 줄들.
    apply_errors: 샵마인 [송장번호수정] 결과 창에 떴던 오류 문구 원문. 어느
    주문인지까지 알아낸 건은 목록에서 '반영오류'로 표시되지만, 문구에서
    주문번호를 못 찾는 경우가 있어 원문 자체도 요약 시트에 그대로 남긴다.
    """
    path = Path(path)
    wb = Workbook()
    _write_entries_sheet(wb.active, entries, applied_label, title)
    _write_summary_sheet(wb.create_sheet(SUMMARY_SHEET_NAME), entries, summary or [],
                         apply_errors or [])
    stale = stale_entries(entries)
    if stale:
        _write_stale_sheet(wb.create_sheet(STALE_SHEET_NAME), stale)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _write_entries_sheet(ws, entries: list[ReportEntry], applied_label: str,
                         title: str) -> None:
    ws.title = SHEET_NAME
    last_col = len(HEADERS)
    rows = sorted(entries, key=_sort_key)

    _write_title(ws, title, rows, last_col)

    ws.append(HEADERS)
    for cell in ws[_HEADER_ROW]:
        cell.font = Font(bold=True, color="202124")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER
    ws.row_dimensions[_HEADER_ROW].height = 22

    for i, entry in enumerate(rows):
        label = result_label(entry)
        ws.append([
            label,
            entry.order_id,
            entry.recipient_name or "",
            order_date_mod.describe(entry.order_date),
            entry.delivery_note or "",
            (entry.product_url or "") if _order_is_old(entry) else "",
            entry.courier or "",
            entry.tracking_no or "",
            _applied_cell(entry, applied_label),
            entry.reason or "",
        ])
        # 덩어리가 바뀌는 자리(그룹 끝)에는 굵은 선을 그어 나눈다. 결과
        # 이름이 아니라 정렬 자리로 비교해야 '주문일지연으로 끌어올린 성공
        # 건'과 '그냥 성공 건'도 선으로 나뉜다.
        next_key = _sort_key(rows[i + 1]) if i + 1 < len(rows) else None
        _style_row(ws[ws.max_row], label, group_end=next_key != _sort_key(entry),
                   stale=is_stale_entry(entry))

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{_HEADER_ROW + 1}"
    ws.auto_filter.ref = (f"A{_HEADER_ROW}:{get_column_letter(last_col)}"
                          f"{max(ws.max_row, _HEADER_ROW)}")


def _order_is_old(entry: ReportEntry) -> bool:
    """주문일이 오늘과 STALE_DAYS 이상 벌어졌는가 (결과가 무엇이든).

    '주문일지연' 시트에 들어가는 기준(report.is_stale_entry)보다 넓다. 그쪽은
    아직 송장이 안 나온 스킵 건만 모으지만, 상품URL은 실패 건처럼 사람이
    화면을 열어봐야 하는 오래된 주문이면 어느 결과든 필요해서다.
    """
    return order_date_mod.is_stale(entry.order_date)


def _applied_cell(entry: ReportEntry, applied_label: str) -> str:
    """'샵마인 반영' 칸 문구. 반영 오류가 난 주문만 따로 표시한다."""
    if entry.category == "apply_error":
        return "반영오류"
    return applied_label if entry.status == "success" else ""


def _write_title(ws, title: str, rows: list[ReportEntry], last_col: int) -> None:
    """맨 위 두 줄: 제목(검은 띠)과 결과별 건수 한 줄."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    head = ws.cell(row=1, column=1, value=f"{title}   전체 {len(rows)}건")
    head.font = Font(bold=True, size=14, color="FFFFFF")
    head.fill = _TITLE_FILL
    head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    counts = "     ".join(f"{label} {n}건" for label, n in label_counts(rows))
    # 주문일이 오래된 건은 결과별 건수와 성격이 달라(성공/스킵에 걸쳐 있다)
    # 같은 줄 끝에 덧붙인다 - 파일을 열자마자 몇 건인지 보이게.
    stale = stale_entries(rows)
    if stale:
        counts += (f"          주문일 {order_date_mod.STALE_DAYS}일 이상 지남 "
                   f"{len(stale)}건 ('{STALE_SHEET_NAME}' 시트 참고)")
    line = ws.cell(row=2, column=1, value=counts)
    line.font = Font(bold=True, size=11, color="3C4043")
    line.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22


def _style_row(row, label: str, *, group_end: bool, stale: bool = False) -> None:
    badge_fill, badge_font, row_fill = _COLORS.get(label, _DEFAULT_COLOR)
    fill = PatternFill("solid", fgColor=row_fill)
    border = Border(left=_THIN, right=_THIN, top=_THIN,
                    bottom=_GROUP_LINE if group_end else _THIN)
    for cell in row:
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")

    badge = row[COL_RESULT]
    badge.fill = PatternFill("solid", fgColor=badge_fill)
    badge.font = Font(bold=True, color=badge_font)
    badge.alignment = Alignment(horizontal="center", vertical="center")

    for cell in (row[COL_ORDER_ID], row[COL_TRACKING]):
        cell.number_format = _TEXT_FORMAT
    row[COL_APPLIED].alignment = Alignment(horizontal="center", vertical="center")
    row[COL_REASON].alignment = Alignment(wrap_text=True, vertical="center")
    row[COL_URL].alignment = Alignment(vertical="center")
    if stale:
        row[COL_ORDER_DATE].font = _STALE_FONT


def _write_stale_sheet(ws, stale: list[ReportEntry]) -> None:
    """주문일이 오래된 스킵 건(취소/품절 제외)만 모은 시트.

    첫 시트에도 같은 건이 들어 있다. 그래도 따로 두는 이유는, 사람이 봐야
    하는 건 '어느 주문이 며칠째 안 나가고 있나'인데 그건 다른 결과들 사이에
    흩어놓으면 한눈에 안 들어와서다.

    공급사 화면에서 읽은 출고/도착 예정 문구와 상품URL을 같이 낸다 - 그 둘만
    있으면 사이트를 다시 찾아 들어가지 않고도 '기다리면 되는 건지'를 판단할
    수 있다.
    """
    _write_summary_band(ws, 1, f"주문일이 {order_date_mod.STALE_DAYS}일 이상 지난 주문 "
                               f"{len(stale)}건 (발송이 늦어지는지 확인해주세요)",
                        last_col=len(STALE_HEADERS))

    ws.append(STALE_HEADERS)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="202124")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER
    ws.row_dimensions[2].height = 22

    for entry in stale:
        label = result_label(entry)
        ws.append([
            entry.order_id,
            entry.recipient_name or "",
            f"{entry.order_date:%Y-%m-%d}",
            f"{order_date_mod.days_since(entry.order_date)}일",
            entry.delivery_note or "",
            entry.product_url or "",
            label,
            entry.reason or "",
        ])
        row = ws[ws.max_row]
        badge_fill, badge_font, row_fill = _COLORS.get(label, _DEFAULT_COLOR)
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.border = _CELL_BORDER
            cell.alignment = Alignment(vertical="center")
        row[0].number_format = _TEXT_FORMAT  # 마켓 주문번호
        # 며칠 지났는지가 이 시트의 핵심이라 제일 눈에 띄게 둔다.
        row[STALE_COL_DAYS].font = _STALE_FONT
        row[STALE_COL_DAYS].alignment = Alignment(horizontal="center", vertical="center")
        row[STALE_COL_RESULT].fill = PatternFill("solid", fgColor=badge_fill)
        row[STALE_COL_RESULT].font = Font(bold=True, color=badge_font)
        row[STALE_COL_RESULT].alignment = Alignment(horizontal="center", vertical="center")
        row[STALE_COL_REASON].alignment = Alignment(wrap_text=True, vertical="center")

    for i, width in enumerate(STALE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A3"


def _write_summary_sheet(ws, entries: list[ReportEntry],
                         summary: list[tuple[str, str]],
                         apply_errors: list[str] | None = None) -> None:
    """요약 시트: 결과별 건수와 할 일을 먼저, 실행 정보/파일 경로를 그 아래에."""
    _write_summary_band(ws, 1, "결과별 건수")
    for col, name in enumerate(("결과", "건수", "해야 할 일"), start=1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER

    for label, n in label_counts(entries):
        ws.append([label, n, _ACTIONS.get(label, "")])
        row = ws[ws.max_row]
        badge_fill, badge_font, row_fill = _COLORS.get(label, _DEFAULT_COLOR)
        for cell in row[:3]:
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.border = _CELL_BORDER
            cell.alignment = Alignment(vertical="center")
        row[0].fill = PatternFill("solid", fgColor=badge_fill)
        row[0].font = Font(bold=True, color=badge_font)
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].font = Font(bold=True)
        row[1].alignment = Alignment(horizontal="center", vertical="center")

    stale = stale_entries(entries)
    if stale:
        _write_summary_band(ws, ws.max_row + 2, "따로 확인할 것")
        ws.append([f"주문일 {order_date_mod.STALE_DAYS}일 이상", len(stale),
                   f"아직 송장이 안 나왔는데 주문한 지 오래된 건 - '{STALE_SHEET_NAME}' 시트에 정리해뒀습니다"])
        row = ws[ws.max_row]
        row[0].font = Font(bold=True)
        row[1].font = _STALE_FONT
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        for cell in row[:3]:
            cell.border = _CELL_BORDER

    if apply_errors:
        _write_summary_band(ws, ws.max_row + 2, "샵마인 [송장번호수정] 결과 창의 오류")
        for line in apply_errors:
            ws.append(["", "", line])
            cell = ws.cell(row=ws.max_row, column=3)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(color="B31412")

    if summary:
        _write_summary_band(ws, ws.max_row + 2, "실행 정보")
        for name, value in summary:
            ws.append([name, value])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=ws.max_row, start_column=2,
                           end_row=ws.max_row, end_column=3)
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(
                wrap_text=True, vertical="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 78


def _write_summary_band(ws, row: int, text: str, *, last_col: int = 3) -> None:
    """요약/지연 시트의 구역 제목(검은 띠) 한 줄."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


def save_run_result(entries, counts: dict, *, applied_label: str, apply_note: str = "",
                    paths=(), apply_errors=(), out_dir: Path | None = None,
                    log=print) -> Path | None:
    """한 번의 실행 결과를 '송장조회결과_시각.xlsx'로 저장한다.

    paths: 요약 시트에 남길 (이름, 경로) 목록. 값이 없는 항목은 뺀다.
    저장에 실패해도 실행 결과 자체를 덮으면 안 되므로, 예외는 경고 한 줄로
    남기고 None을 돌려준다 - 조회 결과는 logs/run_*.json에도 남아 있다.
    """
    if not entries:
        return None
    out_dir = out_dir or DEFAULT_DIR
    now = datetime.now()
    path = out_dir / f"송장조회결과_{now:%Y%m%d_%H%M%S}.xlsx"
    summary = [
        ("실행 시각", f"{now:%Y-%m-%d %H:%M:%S}"),
        ("조회 성공", f"{counts.get('success', 0)}건"),
        ("조회 실패", f"{counts.get('fail', 0)}건"),
        ("조회 스킵", f"{counts.get('skip', 0)}건"),
    ]
    if apply_note:
        summary.append(("샵마인 반영", apply_note))
    summary.extend((name, str(value)) for name, value in paths if value)
    try:
        saved = write_result_excel(entries, path, applied_label=applied_label,
                                   summary=summary, apply_errors=list(apply_errors),
                                   title=f"송장조회 결과  {now:%Y-%m-%d %H:%M}")
    except Exception as e:  # noqa: BLE001 - 결과 저장 실패가 실행 결과를 덮으면 안 된다
        log(f"경고: 결과 엑셀을 저장하지 못했습니다 - {e}")
        return None
    log(f"조회 결과 엑셀: {saved}")
    return saved
